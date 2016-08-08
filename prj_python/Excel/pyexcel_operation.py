# coding:utf-8 
# python3 Anaconda3

import datetime
from openpyxl import Workbook
# from openpyxl.compat import range
from openpyxl.cell import get_column_letter

wb = Workbook()

dest_filename = './empty_book.xlsx'

ws1 = wb.active
ws1.title = "range names"

for row in range(1, 40):
    ws1.append(range(600))

ws1.append([1,2,3])
ws1.append([4,5,6])

for row in ws1.iter_rows('A1:D3'):
    for cell in row:
        print(cell.value)

ws2 = wb.create_sheet(title="Pi")

print(wb.get_sheet_names())


# ws2['F5'] = 3.14

# ws3 = wb.create_sheet(title="Data")
# for row in range(10, 20):
#     for col in range(27, 54):
#         _ = ws3.cell(column=col, row=row, value="%s" % get_column_letter(col))
# print(ws3['AA10'].value)

wb.save(filename = dest_filename)

wb = Workbook()
ws = wb.active

data = [
    ["Fruit", "Quantity"],
    ["Kiwi", 3],
    ["Grape", 15],
    ["Apple", 3],
    ["Peach", 3],
    ["Pomegranate", 3],
    ["Pear", 3],
    ["Tangerine", 3],
    ["Blueberry", 3],
    ["Mango", 3],
    ["Watermelon", 3],
    ["Blackberry", 3],
    ["Orange", 3],
    ["Raspberry", 3],
    ["Banana", 3]
]

for r in data:
    ws.append(r)

ws.auto_filter.ref = "A1:B15"
ws.auto_filter.add_filter_column(0, ["Kiwi", "Apple", "Mango"])
ws.auto_filter.add_sort_condition("B2:B15")

wb.save("filtered.xlsx")
